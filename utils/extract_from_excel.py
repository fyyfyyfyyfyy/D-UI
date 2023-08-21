import copy
import json
import os
from typing import List, Union

from openpyxl import load_workbook  # type: ignore

json_model = {}

with open(r"./example/example.json", 'r', encoding='utf-8') as json_file:
    json_model = json.load(json_file)


def load_from_excel(
    file: str = './example/example_dirty.xlsx',
    write_to_file: bool = False,
    output_file: str = "example_out_clean.json"
) -> List[dict]:

    wb = load_workbook(file, data_only=True)

    sh = wb.worksheets[0]   # Todo: 添加多表支持，目前默认读取第一个表格

    json_output = []

    start_row, start_col = find_data_region(sh)

    if not start_row:
        print("未检测到数据，检查表中是否有“事件内容分解”单元格")
        # 目前检测数据区域依靠表头的“事件内容分解”

    print(f"检测到的数据区域：{(start_col, start_row, start_col + 58, sh.max_row + 1)}")

    for row_index in range(start_row, sh.max_row + 1):
        data: List[Union[str, None]] = []
        for col_index in range(start_col, start_col + 58):
            r = sh.cell(row=row_index, column=col_index)
            if r.value is None or r.value == "\\" or r.value == "/":
                data.append(None)
            else:
                data.append(r.value)

        json_input = copy.deepcopy(json_model)

        fill_in(json_input, data)

        remove_empty_fields(json_input)

        if json_input == {}:
            continue

        json_output.append(json_input)

    print(f"共有{len(json_output)}条数据")

    if write_to_file:
        with open(output_file, 'w', encoding="utf-8") as json_file:
            json_formatted = json.dumps(json_output, indent=4, ensure_ascii=False)
            json_file.write(json_formatted)
        print(f"输出到文件{os.path.abspath(output_file)}")

    return json_output


def find_data_region(sheet):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == "事件内容分解":
                return cell.row + 2, cell.column
    return None, None


def remove_empty_fields(data):
    keys_to_remove = []

    for key, value in data.items():
        if isinstance(value, dict):
            remove_empty_fields(value)
            if not value:
                keys_to_remove.append(key)
        elif value in [None]:
            keys_to_remove.append(key)

    for key in keys_to_remove:
        data.pop(key, None)


def fill_in(json_input, rows):
    json_input["背景事件内容"]["输入背景内容"]["事件内容分解"]["内容"] = rows[0]
    json_input["背景事件内容"]["记录事件"]["执行动作"]["人工/GPT推理"] = rows[7]
    json_input["背景事件内容"]["记录事件"]["事件时间"]["年"] = rows[8]
    json_input["背景事件内容"]["记录事件"]["事件时间"]["月"] = rows[9]
    json_input["背景事件内容"]["记录事件"]["事件时间"]["日"] = rows[10]
    json_input["背景事件内容"]["记录事件"]["事件时间"]["时"] = rows[11]
    json_input["背景事件内容"]["记录事件"]["事件时间"]["分"] = rows[12]
    json_input["背景事件内容"]["记录事件"]["事件时间"]["秒"] = rows[13]
    json_input["背景事件内容"]["记录事件"]["事件地点"]["国家"] = rows[14]
    json_input["背景事件内容"]["记录事件"]["事件地点"]["省"] = rows[15]
    json_input["背景事件内容"]["记录事件"]["事件地点"]["市"] = rows[16]
    json_input["背景事件内容"]["记录事件"]["事件地点"]["县"] = rows[17]
    json_input["背景事件内容"]["记录事件"]["事件地点"]["街道"] = rows[18]
    json_input["背景事件内容"]["记录事件"]["事件地点"]["门牌号"] = rows[19]
    json_input["背景事件内容"]["记录事件"]["事件地点"]["地点文本"] = rows[20]
    json_input["背景事件内容"]["记录事件"]["事件信念"]["引导语"] = rows[21]
    json_input["背景事件内容"]["记录事件"]["事件信念"]["欲望"] = rows[22]
    json_input["背景事件内容"]["记录事件"]["事件信念"]["中间词"] = rows[23]
    json_input["背景事件内容"]["记录事件"]["事件信念"]["信念核心"] = rows[24]
    json_input["背景事件内容"]["记录事件"]["事件信念"]["信念描述（标准语句）"] = rows[25]
    json_input["背景事件内容"]["记录事件"]["GPT推理依据"]["推理关键词"] = rows[26]
    json_input["背景事件内容"]["记录事件"]["GPT推理依据"]["推理过程"] = rows[27]
    json_input["背景事件内容"]["记录事件"]["欲望数据"]["人物"] = rows[28]
    json_input["背景事件内容"]["记录事件"]["欲望数据"]["L3欲望编号"] = rows[29]
    json_input["背景事件内容"]["记录事件"]["欲望数据"]["L3欲望"] = rows[30]
    json_input["背景事件内容"]["记录事件"]["欲望数据"]["L4欲望"] = rows[31]
    json_input["背景事件内容"]["记录事件"]["欲望数据"]["L5欲望"] = rows[32]
    json_input["背景事件内容"]["记录事件"]["欲望数据"]["欲望值"] = rows[33]
    json_input["背景事件内容"]["记录事件"]["欲望数据"]["识别欲望的关键词"] = rows[34]
    json_input["背景事件内容"]["记录事件"]["事件关联感受数值"]["FH（开心）"] = rows[35]
    json_input["背景事件内容"]["记录事件"]["事件关联感受数值"]["FS（难受）"] = rows[36]
    json_input["背景事件内容"]["记录事件"]["事件关联感受数值"]["FU（讨厌）"] = rows[37]
    json_input["背景事件内容"]["记录事件"]["事件关联感受数值"]["FW（惊讶）"] = rows[38]
    json_input["背景事件内容"]["记录事件"]["事件关联感受数值"]["FA（生气）"] = rows[39]
    json_input["分析推理"]["外部信念与感受作用关系"] = rows[40]
    json_input["分析推理"]["生成感受数值"]["FH（开心）"] = rows[41]
    json_input["分析推理"]["生成感受数值"]["FS（难受）"] = rows[42]
    json_input["分析推理"]["生成感受数值"]["FU（讨厌）"] = rows[43]
    json_input["分析推理"]["生成感受数值"]["FW（惊讶）"] = rows[44]
    json_input["分析推理"]["生成感受数值"]["FA（生气）"] = rows[45]
    json_input["分析推理"]["情绪计算公式"] = rows[46]
    json_input["输出内容"]["生成情绪数值"]["EH（开心）"] = rows[47]
    json_input["输出内容"]["生成情绪数值"]["ES（难过）"] = rows[48]
    json_input["输出内容"]["生成情绪数值"]["EU（讨厌）"] = rows[49]
    json_input["输出内容"]["生成情绪数值"]["EW（惊讶）"] = rows[50]
    json_input["输出内容"]["生成情绪数值"]["EA（生气）"] = rows[51]
    json_input["输出内容"]["输出信念Rx"]["正向信念"] = rows[52]
    json_input["输出内容"]["输出信念Rx"]["负向信念"] = rows[53]
    json_input["输出内容"]["表达习惯（HL）"]["开头语"] = rows[54]
    json_input["输出内容"]["表达习惯（HL）"]["句中常用语"] = rows[55]
    json_input["输出内容"]["表达习惯（HL）"]["结束语气词"] = rows[56]
    json_input["输出内容"]["GPT推理格式"]["格式"] = rows[57]


if __name__ == "__main__":
    load_from_excel()

"""
结构：

背景事件内容
    输入背景内容
        事件内容分解
            内容
    记录事件
        执行动作
            人工/GPT推理
        事件时间
            年
            月
            日
            时
            分
            秒
        事件地点
            国家
            省
            市
            县
            街道
            门牌号
            地点文本
        事件信念
            引导语
            欲望
            中间词
            信念核心
            信念描述（标准语句）
        GPT推理依据
            推理关键词
            推理过程
        欲望数据
            人物
            L3欲望编号
            L3欲望
            L4欲望
            L5欲望
            欲望值
            识别欲望的关键词
        事件关联感受数值
            FH（开心）
            FS（难受）
            FU（讨厌）
            FW（惊讶）
            FA（生气）
分析推理
    外部信念与感受作用关系
    生成感受数值
        FH（开心）
        FS（难受）
        FU（讨厌）
        FW（惊讶）
        FA（生气）
    情绪计算公式

输出内容
    生成情绪数值
        EH（开心）
        ES（难过）
        EU（讨厌）
        EW(惊讶）
        EA（生气）
    输出信念Rx
        正向信念
        负向信念
    表达习惯（HL）
        开头语
        句中常用语
        结束语气词
    GPT推理格式
        格式

展平后的数据与对应表格的列数：

0   内容
7   人工/GPT推理
8   年
9   月
10  日
11  时
12  分
13  秒
14  国家
15  省
16  市
17  县
18  街道
19  门牌号
20  地点文本
21  引导语
22  欲望
23  中间词
24  信念核心
25  信念描述（标准语句）
26  推理关键词
27  推理过程
28  人物
29  L3欲望编号
30  L3欲望
31  L4欲望
32  L5欲望
33  欲望值
34  识别欲望的关键词
35  FH（开心）
36  FS（难受）
37  FU（讨厌）
38  FW（惊讶）
39  FA（生气）
40  外部信念与感受作用关系
41  FH（开心）
42  FS（难受）
43  FU（讨厌）
44  FW（惊讶）
45  FA（生气）
46  情绪计算公式
47  EH（开心）
48  ES（难过）
49  EU（讨厌）
50  EW（惊讶）
51  EA（生气）
52  正向信念
53  负向信念
54  开头语
55  句中常用语
56  结束语气词
57  格式

"""
